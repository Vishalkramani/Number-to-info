import os
import time
import sys
import subprocess

# Function to check and install modules
def install_and_import(package):
    try:
        __import__(package)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        __import__(package)

# Install necessary modules
modules = ['colorama', 'openpyxl']
for module in modules:
    install_and_import(module)

# Clear the terminal after installing modules
os.system('cls' if os.name == 'nt' else 'clear')

# Import the modules after ensuring they are installed
from colorama import init, Fore, Style
from openpyxl import load_workbook

# Initialize colorama for Windows compatibility
init(autoreset=True)

# ASCII art banner with animation
banner = """
 _ _   _ _  
|_   _| \ | |_   _| 
  | | |  \| | | |   
  | | | . ` | | |   
  | | | |\  | | |   
  \_/ \_| \_/ \_/
"""
def display_banner():
    for line in banner.splitlines():
        print(Fore.GREEN + line)
        time.sleep(0.2)  # Animation effect
    time.sleep(1)

# Load the Excel file and check the mobile number manually
def check_mobile_number(mobile_number):
    try:
        # Load the Excel file
        workbook = load_workbook('Govt.xlsx')
        sheet = workbook.active

        # Get the header to identify columns
        header = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}

        # Search for the mobile number in the 'Mobile' column
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[header['Mobile'] - 1]) == mobile_number:
                print(Fore.YELLOW + "\nDetails found for the mobile number:")
                details = {
                    'Name': row[header['Name'] - 1],
                    'Mobile': row[header['Mobile'] - 1],
                    'Email': row[header['Email'] - 1],
                    'Address': row[header['Address'] - 1],
                    'City': row[header['City'] - 1],
                    'State': row[header['State'] - 1],
                    'Industry': row[header['Industry'] - 1],
                    'DOB': row[header['Dob'] - 1]
                }
                for key, value in details.items():
                    print(Fore.CYAN + f"{key}: {value}")
                return
        print(Fore.RED + "No details found for this mobile number.")
    except FileNotFoundError:
        print(Fore.RED + "Govt.xlsx file not found. Please make sure the file exists.")
    except Exception as e:
        print(Fore.RED + f"An error occurred: {e}")

# Main function
def main():
    # Display the banner with animation
    display_banner()

    # Input mobile number
    print(Fore.CYAN + "\nPlease enter the mobile number in the format XXXXXXXXXX")
    mobile_number = input(Fore.YELLOW + "Mobile Number: ").strip()

    # Check the format of the mobile number
    if len(mobile_number) == 10 and mobile_number.isdigit():
        # Check if mobile number exists in the Govt.xlsx file
        check_mobile_number(mobile_number)
    else:
        print(Fore.RED + "Invalid mobile number format. Please enter a 10-digit number.")

if __name__ == "__main__":
    main()